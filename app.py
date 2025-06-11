#-------------------------------------------------------------------------------------------------------------------------
# Para rodar: no terminal digite
# python -m streamlit run app.py
#-------------------------------------------------------------------------------------------------------------------------

import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

# Configuração da página
st.set_page_config(page_title="Analisador de Vendas", layout="centered")

# Cabeçalho com logo
col1, col2, col3 = st.columns([1, 3, 1])

st.title("Analisador de Vendas com Python")
st.write("Faça upload da planilha de vendas (.xlsx) e gere um relatório profissional em Excel com resumos visuais.")

# Upload do arquivo
uploaded_file = st.file_uploader("Faça upload da planilha de vendas", type=["xlsx"])

# Função para gerar Excel profissional
def gerar_excel(df):
    output = BytesIO()

    # Preparar dados
    df["Mês"] = pd.to_datetime(df["Data da Venda"]).dt.to_period("M").astype(str)
    total_geral = df["Valor da Venda"].sum()

    produtos_valor = df.groupby("Produto")["Valor da Venda"].sum().reset_index()
    produtos_qtd = df["Produto"].value_counts().reset_index()
    produtos_qtd.columns = ["Produto", "Quantidade"]
    produtos_final = pd.merge(produtos_valor, produtos_qtd, on="Produto")
    produtos_final["Ticket Médio"] = produtos_final["Valor da Venda"] / produtos_final["Quantidade"]

    meses = df.groupby("Mês")["Valor da Venda"].sum().reset_index()
    regioes = df.groupby("Região")["Valor da Venda"].sum().reset_index()
    lista_produtos = pd.DataFrame(sorted(df["Produto"].unique()), columns=["Produtos do Supermercado"])

    # Estilo
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    bold_font = Font(bold=True)
    centered = Alignment(horizontal="center")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Criar workbook
    wb = Workbook()
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    ws_resumo.append(["Resumo das Vendas"])
    ws_resumo.append(["Total Geral Vendido", f"R$ {total_geral:,.2f}"])

    # Por Produto
    ws_prod = wb.create_sheet("Por Produto")
    ws_prod.append(["Produto", "Valor da Venda (R$)", "Quantidade de Vendas", "Ticket Médio (R$)"])
    for _, row in produtos_final.iterrows():
        ws_prod.append([
            row["Produto"], row["Valor da Venda"], row["Quantidade"], round(row["Ticket Médio"], 2)
        ])

    # Por Mês
    ws_mes = wb.create_sheet("Por Mês")
    ws_mes.append(["Mês", "Total Vendido (R$)"])
    for _, row in meses.iterrows():
        ws_mes.append([row["Mês"], row["Valor da Venda"]])

    # Por Região
    ws_reg = wb.create_sheet("Por Região")
    ws_reg.append(["Região", "Total Vendido (R$)"])
    for _, row in regioes.iterrows():
        ws_reg.append([row["Região"], row["Valor da Venda"]])

    # Lista de Produtos
    ws_lista = wb.create_sheet("Lista de Produtos")
    ws_lista.append(["Produtos do Supermercado"])
    for produto in lista_produtos["Produtos do Supermercado"]:
        ws_lista.append([produto])

    # Formatação
    for ws in [ws_prod, ws_mes, ws_reg, ws_lista]:
        for cell in ws[1]:
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = centered
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border

    # Gráfico na aba "Por Produto"
    chart = BarChart()
    chart.title = "Faturamento por Produto"
    chart.y_axis.title = "R$"
    chart.x_axis.title = "Produto"
    data = Reference(ws_prod, min_col=2, min_row=1, max_row=ws_prod.max_row, max_col=2)
    cats = Reference(ws_prod, min_col=1, min_row=2, max_row=ws_prod.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    ws_prod.add_chart(chart, "F25")

    wb.save(output)
    return output.getvalue()

# Execução principal
if uploaded_file:
    df = pd.read_excel(uploaded_file)

    colunas_esperadas = {"Data da Venda", "Produto", "Região", "Valor da Venda"}
    if not colunas_esperadas.issubset(set(df.columns)):
        st.error("A planilha deve conter as colunas: Data da Venda, Produto, Região, Valor da Venda.")
    else:
        st.success("Planilha validada com sucesso. Pronto para gerar o relatório!")

        #Análises para a interface
        df["Data da Venda"] = pd.to_datetime(df["Data da Venda"])
        df["Mês"] = df["Data da Venda"].dt.to_period("M").astype(str)
        total_geral = df["Valor da Venda"].sum()

        produto_valor = df.groupby("Produto")["Valor da Venda"].sum().reset_index()
        produto_faturamento = produto_valor.sort_values("Valor da Venda", ascending=False).iloc[0]

        produto_qtd = df["Produto"].value_counts()
        produto_top_qtd = produto_qtd.idxmax()
        qtd_vendas = produto_qtd.max()

        regiao_top = df.groupby("Região")["Valor da Venda"].sum().reset_index().sort_values("Valor da Venda", ascending=False).iloc[0]

        #Exibição no site
        st.subheader("Resumo das Vendas")
        st.metric(label="Total Geral Vendido", value=f"R$ {total_geral:,.2f}")
        st.success(f"Produto com maior faturamento: {produto_faturamento['Produto']} (R$ {produto_faturamento['Valor da Venda']:,.2f})")
        st.success(f"Produto mais vendido (quantidade): {produto_top_qtd} ({qtd_vendas} vendas)")
        st.success(f"Região com maior faturamento: {regiao_top['Região']}")

        # Botão de download
        st.download_button(
            label="Baixar Relatório Excel",
            data=gerar_excel(df),
            file_name=f"relatorio_vendas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
